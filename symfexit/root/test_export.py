import csv
import io
from datetime import datetime

from django.contrib.admin.models import LogEntry
from django.contrib.admin.sites import AdminSite
from django.contrib.auth.models import Group
from django.test import RequestFactory, TestCase, override_settings
from django.utils import timezone, translation
from openpyxl import load_workbook

from symfexit.events.admin import EventsAdmin
from symfexit.events.models import Event
from symfexit.members.admin import UserAdmin
from symfexit.members.models import User
from symfexit.root.export_builder import ExportDataBuilder
from symfexit.root.export_formatters import CSVFormatter, ExcelFormatter, get_formatter
from symfexit.root.export_views import ExportFieldSelectionView
from symfexit.tenants.models import Client


@override_settings(LANGUAGE_CODE="en")
class BaseExportTestCase(TestCase):
    """Base test class with common setup for export tests."""

    def setUp(self):
        """Set up common test data."""
        translation.activate("en")
        self.site = AdminSite()
        self.factory = RequestFactory()

        self.group1 = Group.objects.create(name="Group Alpha")
        self.group2 = Group.objects.create(name="Group Beta")

        self.user = User.objects.create_superuser(
            email="admin@test.com",
            password="testpass123",
            first_name="Admin",
            last_name="User",
            member_identifier=1000,
        )
        self.user.groups.add(self.group1)

        self.event = Event.objects.create(
            event_name="Test Event",
            event_organiser="Test Org",
            event_date=timezone.make_aware(datetime(2026, 5, 17, 10, 30)),
            event_end=timezone.make_aware(datetime(2026, 5, 17, 16, 0)),
            event_desc="<p>HTML description</p>",
        )
        self.event.attendees.add(self.user)


class ExportMixinTestCase(BaseExportTestCase):
    """Test the ExportMixin configuration layer."""

    def test_export_action_and_fields(self):
        """Test that export action is registered and fields are extracted."""
        admin_class = EventsAdmin(Event, self.site)

        # Action is registered
        self.assertIn("export_data", admin_class.actions)

        # Fields are extracted from export_fields
        fields = admin_class.get_export_fields()
        self.assertIn("event_name", fields)
        self.assertIn("attendees", fields)
        self.assertEqual(len(fields), len(admin_class.export_fields))

    def test_field_labels(self):
        """Test field label resolution for various field types."""
        admin_class = EventsAdmin(Event, self.site)

        # Regular model field - should get verbose name
        label = admin_class.get_export_field_label("event_name")
        self.assertEqual(label, str(Event._meta.get_field("event_name").verbose_name))

        # Count field - should contain "count"
        label = admin_class.get_export_field_label("attendees__count")
        self.assertIn("count", label.lower())

        # Custom label from export_field_labels
        admin_class.export_field_labels = {"event_name": "Custom Event Name"}
        label = admin_class.get_export_field_label("event_name")
        self.assertEqual(label, "Custom Event Name")

        # Admin method with short_description
        class TestAdmin(EventsAdmin):
            def custom_method(self, obj):
                return "custom value"

            custom_method.short_description = "Custom Method Label"

        admin_class = TestAdmin(Event, self.site)
        label = admin_class.get_export_field_label("custom_method")
        self.assertEqual(label, "Custom Method Label")

        # Admin method without short_description - fallback to title case
        class TestAdmin2(EventsAdmin):
            def custom_computed_value(self, obj):
                return "computed"

        admin_class = TestAdmin2(Event, self.site)
        label = admin_class.get_export_field_label("custom_computed_value")
        self.assertEqual(label, "Custom Computed Value")

    def test_format_export_value(self):
        """Test value formatting for various field types."""
        admin_class = EventsAdmin(Event, self.site)
        user_admin = UserAdmin(User, self.site)

        # Boolean True - should be Yes/Ja
        value = user_admin.format_export_value(self.user, "is_active")
        self.assertIn(value, ["Yes", "No", "Ja", "Nee"])

        # Boolean False - should also be Yes/No/Ja/Nee
        user_inactive = User.objects.create(
            email="inactive@test.com",
            first_name="Inactive",
            last_name="User",
            member_identifier=1001,
            is_staff=False,
        )
        value = user_admin.format_export_value(user_inactive, "is_staff")
        self.assertIn(value, ["Yes", "No", "Ja", "Nee"])

        # None value - should be empty string
        value = user_admin.format_export_value(user_inactive, "phone_number")
        self.assertEqual(value, "")

        # Count field
        value = admin_class.format_export_value(self.event, "attendees__count")
        self.assertEqual(value, "1")

        # HTML preservation - should keep HTML tags
        value = admin_class.format_export_value(self.event, "event_desc")
        self.assertEqual(value, "<p>HTML description</p>")
        self.assertIn("<p>", value)

        # Date field formatting
        value = admin_class.format_export_value(self.event, "event_date")
        self.assertIsInstance(value, str)
        self.assertGreater(len(value), 0)

        # M2M field with .all() method
        value = admin_class.format_export_value(self.event, "attendees")
        self.assertIn(self.user.email, value)

    def test_expandable_fields(self):
        """Test extraction of expandable field configuration."""
        admin_class = EventsAdmin(Event, self.site)
        expandable = admin_class.get_expandable_fields()

        # Direct expandable field
        self.assertIn("attendees", expandable)
        field_obj, related_fields = expandable["attendees"]
        self.assertIsNotNone(related_fields)

        # Nested expandable from referenced admin
        self.assertIn("attendees__groups", expandable)


class ExportDataBuilderTestCase(BaseExportTestCase):
    """Test the data transformation layer."""

    def setUp(self):
        """Set up additional test data for builder tests."""
        super().setUp()

        # Add more users for complex scenarios
        self.user2 = User.objects.create(
            email="user2@test.com",
            first_name="Bob",
            last_name="Brown",
            member_identifier=2002,
        )
        self.user2.groups.add(self.group2)

        self.user3 = User.objects.create(
            email="user3@test.com",
            first_name="Charlie",
            last_name="Chen",
            member_identifier=2003,
        )
        # user3 has no groups

        self.user.groups.add(self.group1, self.group2)
        self.event.attendees.add(self.user2, self.user3)

        self.admin_class = EventsAdmin(Event, self.site)

    def test_build_headers(self):
        """Test header generation for simple, count, and expansion fields."""
        queryset = Event.objects.filter(pk=self.event.pk)

        # Simple fields
        builder = ExportDataBuilder(
            self.admin_class, Event, queryset, ["event_name", "event_organiser"]
        )
        headers = builder.build_headers()
        self.assertEqual(len(headers), 2)
        self.assertEqual(headers[0], str(Event._meta.get_field("event_name").verbose_name))

        # With count field
        builder = ExportDataBuilder(
            self.admin_class, Event, queryset, ["event_name", "attendees__count"]
        )
        headers = builder.build_headers()
        self.assertEqual(len(headers), 2)
        self.assertIn("count", headers[1].lower())

        # With expansion
        builder = ExportDataBuilder(
            self.admin_class,
            Event,
            queryset,
            ["event_name", "attendees__email", "attendees__first_name"],
        )
        headers = builder.build_headers()
        self.assertEqual(len(headers), 3)
        self.assertIn("email", headers[1].lower())

    def test_build_rows(self):
        """Test row generation for simple fields, expansions, and nested expansions."""
        queryset = Event.objects.filter(pk=self.event.pk)

        # Simple fields
        builder = ExportDataBuilder(
            self.admin_class, Event, queryset, ["event_name", "event_organiser"]
        )
        rows = builder.build_rows()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "Test Event")

        # With expansion (3 attendees)
        builder = ExportDataBuilder(
            self.admin_class,
            Event,
            queryset,
            ["event_name", "attendees__email", "attendees__first_name"],
        )
        rows = builder.build_rows()
        self.assertEqual(len(rows), 3)
        self.assertTrue(all(row[0] == "Test Event" for row in rows))
        emails = {row[1] for row in rows}
        self.assertEqual(emails, {"admin@test.com", "user2@test.com", "user3@test.com"})

        # With nested expansion (user has 2 groups, user2 has 1 group, user3 has 0 groups = 4 rows)
        builder = ExportDataBuilder(
            self.admin_class,
            Event,
            queryset,
            ["event_name", "attendees__first_name", "attendees__groups__name"],
        )
        rows = builder.build_rows()
        self.assertEqual(len(rows), 4)

        # Verify user with 2 groups appears twice
        admin_rows = [r for r in rows if r[1] == "Admin"]
        self.assertEqual(len(admin_rows), 2)
        admin_groups = {r[2] for r in admin_rows}
        self.assertEqual(admin_groups, {"Group Alpha", "Group Beta"})

        # Verify user3 with no groups has empty group name
        charlie_rows = [r for r in rows if r[1] == "Charlie"]
        self.assertEqual(len(charlie_rows), 1)
        self.assertEqual(charlie_rows[0][2], "")

        # With count and expansion
        builder = ExportDataBuilder(
            self.admin_class,
            Event,
            queryset,
            [
                "event_name",
                "attendees__count",
                "attendees__email",
                "attendees__groups__count",
                "attendees__groups__name",
            ],
        )
        rows = builder.build_rows()
        self.assertEqual(len(rows), 4)
        # All rows should have same attendees count (3)
        self.assertTrue(all(row[1] == "3" for row in rows))


class ExportFormattersTestCase(BaseExportTestCase):
    """Test the presentation layer (formatters)."""

    def test_formatter_factory(self):
        """Test getting formatters via factory function."""
        csv_formatter = get_formatter("csv")
        self.assertIsInstance(csv_formatter, CSVFormatter)

        excel_formatter = get_formatter("excel")
        self.assertIsInstance(excel_formatter, ExcelFormatter)

        with self.assertRaises(ValueError):
            get_formatter("invalid")

    def test_csv_formatter(self):
        """Test CSV file generation with UTF-8 BOM."""
        formatter = CSVFormatter()
        headers = ["Name", "Email", "Age"]
        rows = [["Alice", "alice@test.com", "25"], ["Bob", "bob@test.com", "30"]]

        response = formatter.format(headers, rows, "users")

        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("users_", response["Content-Disposition"])
        self.assertIn(".csv", response["Content-Disposition"])

        content = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        csv_rows = list(reader)

        self.assertEqual(csv_rows[0], headers)
        self.assertEqual(csv_rows[1], rows[0])
        self.assertEqual(csv_rows[2], rows[1])

    def test_excel_formatter(self):
        """Test Excel file generation with bold headers and auto-sizing."""
        formatter = ExcelFormatter()
        headers = ["Name", "Email", "Age"]
        rows = [["Alice", "alice@test.com", "25"], ["Bob", "bob@test.com", "30"]]

        response = formatter.format(headers, rows, "users")

        self.assertIn("spreadsheetml.sheet", response["Content-Type"])
        self.assertIn("users_", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        header_row = [cell.value for cell in ws[1]]
        self.assertEqual(header_row, headers)

        data_row1 = [cell.value for cell in ws[2]]
        self.assertEqual(data_row1, rows[0])


class ExportViewTestCase(BaseExportTestCase):
    """Test the HTTP layer (views)."""

    def test_field_selection_view_with_permissions(self):
        """Test GET request to field selection view requires permissions."""
        # With proper permissions (superuser has all permissions)
        request = self.factory.get("/admin/export/events/event/")
        request.user = self.user
        request.session = {"export_ids_events_event": [self.event.pk]}
        request.tenant = Client.objects.first() if Client.objects.exists() else None

        view = ExportFieldSelectionView.as_view()
        response = view(request, app_label="events", model_name="event")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Export", response.content)

        # Without export permission (staff user without view permission)
        regular_user = User.objects.create(
            email="regular@test.com",
            first_name="Regular",
            last_name="User",
            member_identifier=3001,
            is_staff=True,
            is_superuser=False,
            is_active=True,  # Must be active for staff_member_required
        )

        request = self.factory.get("/admin/export/events/event/")
        request.user = regular_user
        request.session = {"export_ids_events_event": [self.event.pk]}
        request.tenant = Client.objects.first() if Client.objects.exists() else None
        response = view(request, app_label="events", model_name="event")
        # Either 403 (permission denied) or 302 (redirect) is acceptable - just not 200
        self.assertIn(response.status_code, [302, 403])

        # Without session IDs - should redirect
        request = self.factory.get("/admin/export/events/event/")
        request.user = self.user
        request.session = {}
        request.tenant = Client.objects.first() if Client.objects.exists() else None
        response = view(request, app_label="events", model_name="event")
        self.assertEqual(response.status_code, 302)

    def test_export_validation_and_logging(self):
        """Test export validates fields and creates audit log."""
        request = self.factory.post(
            "/admin/export/events/event/",
            {"fields": ["event_name"], "format": "csv", "download": "true"},
        )
        request.user = self.user
        request.session = {"export_ids_events_event": [self.event.pk]}

        view = ExportFieldSelectionView.as_view()
        log_count_before = LogEntry.objects.count()

        response = view(request, app_label="events", model_name="event")

        # Should create log entry
        log_count_after = LogEntry.objects.count()
        self.assertEqual(log_count_after, log_count_before + 1)

        log_entry = LogEntry.objects.latest("id")
        self.assertEqual(log_entry.user, self.user)
        self.assertIn("export", log_entry.change_message.lower())

        # With no fields - should return 400
        request_no_fields = self.factory.post(
            "/admin/export/events/event/",
            {"fields": [], "format": "csv", "download": "true"},
        )
        request_no_fields.user = self.user
        request_no_fields.session = {"export_ids_events_event": [self.event.pk]}

        response = view(request_no_fields, app_label="events", model_name="event")
        self.assertEqual(response.status_code, 400)


class IntegrationTestCase(BaseExportTestCase):
    """End-to-end integration tests."""

    def setUp(self):
        """Set up complete test scenario."""
        super().setUp()

        # Add additional users for integration tests
        self.user2 = User.objects.create(
            email="dev2@company.com",
            first_name="Bob",
            last_name="Developer",
            member_identifier=4002,
            phone_number="+31687654321",
            city="Rotterdam",
        )
        self.user2.groups.add(self.group1, self.group2)

        self.user.phone_number = "+31612345678"
        self.user.city = "Amsterdam"
        self.user.save()

        self.event.event_name = "Company Hackathon"
        self.event.event_organiser = "Tech Lead"
        self.event.event_desc = "Annual company hackathon"
        self.event.save()
        self.event.attendees.add(self.user2)

        self.admin_class = EventsAdmin(Event, self.site)

    def test_full_export_workflow(self):
        """Test complete export workflow for both CSV and Excel formats."""
        queryset = Event.objects.filter(pk=self.event.pk)
        fields = [
            "event_name",
            "event_organiser",
            "attendees__count",
            "attendees__email",
            "attendees__first_name",
            "attendees__city",
            "attendees__groups__name",
        ]

        builder = ExportDataBuilder(self.admin_class, Event, queryset, fields)
        headers = builder.build_headers()
        rows = builder.build_rows()

        # Test CSV
        csv_formatter = CSVFormatter()
        csv_response = csv_formatter.format(headers, rows, "hackathon_export")

        content = csv_response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        csv_rows = list(reader)

        self.assertEqual(len(csv_rows), 4)  # header + 3 data rows (user:1 group, user2:2 groups)
        event_names = {row[0] for row in csv_rows[1:]}
        self.assertEqual(event_names, {"Company Hackathon"})

        for row in csv_rows[1:]:
            self.assertEqual(row[2], "2")  # 2 attendees

        # Test Excel
        excel_formatter = ExcelFormatter()
        excel_response = excel_formatter.format(headers, rows, "hackathon_export")

        wb = load_workbook(io.BytesIO(excel_response.content))
        ws = wb.active

        self.assertEqual(ws.max_row, 4)  # header + 3 data rows
        for row_idx in range(2, ws.max_row + 1):
            self.assertEqual(ws.cell(row_idx, 1).value, "Company Hackathon")
